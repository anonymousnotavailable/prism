"""Tests for modules.data_engine's large-file ingestion path.

load_data()'s existing behavior (encoding/delimiter sniffing, header
recovery, MAX_ROWS truncation) reads the *entire* file into pandas before
deciding what to keep — fine for small files, wasteful and eventually
impossible for a genuinely large one (a multi-GB CSV shouldn't have to
fully materialize in Python memory just to get truncated down to 50k rows
a moment later). This module adds an optional DuckDB-backed path: for
files above a size threshold, DuckDB's out-of-core CSV reader counts rows
and pulls a random sample directly, without pandas ever seeing the full
file. Below the threshold, behavior is byte-for-byte unchanged — these
tests exist to prove that boundary, not to re-test load_data()'s existing
encoding/header logic (already exercised manually across every prior run's
audit).
"""
from __future__ import annotations

import io

import pandas as pd
import pytest

from modules.data_engine import (
    LARGE_FILE_THRESHOLD_BYTES,
    _duckdb_sample_csv,
    _should_attempt_duckdb,
    _should_attempt_streaming_excel,
    _stream_sample_excel,
    check_sampling_fidelity,
    load_data,
)


class _FakeUploadedFile(io.BytesIO):
    """Minimal stand-in for streamlit's UploadedFile — a BytesIO with the
    extra `.name`/`.size` attributes load_data() and its helpers read."""

    def __init__(self, data: bytes, name: str = "data.csv", size: int | None = None):
        super().__init__(data)
        self.name = name
        self.size = size if size is not None else len(data)


def _csv_bytes(n_rows: int) -> bytes:
    lines = ["id,value,label"] + [f"{i},{i * 1.5},group{i % 3}" for i in range(n_rows)]
    return ("\n".join(lines)).encode("utf-8")


pytest.importorskip("duckdb")


# ─────────────────────────────────────────────────────────────────────────
# _should_attempt_duckdb — the size gate
# ─────────────────────────────────────────────────────────────────────────
def test_should_attempt_duckdb_false_for_small_file():
    f = _FakeUploadedFile(_csv_bytes(10), size=1_000)
    assert _should_attempt_duckdb(f) is False


def test_should_attempt_duckdb_true_above_threshold():
    f = _FakeUploadedFile(_csv_bytes(10), size=LARGE_FILE_THRESHOLD_BYTES + 1)
    assert _should_attempt_duckdb(f) is True


def test_should_attempt_duckdb_false_when_size_unknown():
    f = _FakeUploadedFile(_csv_bytes(10))
    del f.size  # simulate an object that doesn't expose .size at all
    assert _should_attempt_duckdb(f) is False


# ─────────────────────────────────────────────────────────────────────────
# _duckdb_sample_csv — the actual out-of-core read
# ─────────────────────────────────────────────────────────────────────────
def test_duckdb_sample_returns_all_rows_when_under_cap():
    f = _FakeUploadedFile(_csv_bytes(50))
    result = _duckdb_sample_csv(f, max_rows=200)
    assert result is not None
    df, warnings = result
    assert len(df) == 50
    assert list(df.columns) == ["id", "value", "label"]
    assert warnings == []


def test_duckdb_sample_takes_random_sample_when_over_cap():
    f = _FakeUploadedFile(_csv_bytes(2000))
    result = _duckdb_sample_csv(f, max_rows=100, random_state=42)
    assert result is not None
    df, warnings = result
    assert len(df) == 100
    assert warnings and "2,000" in warnings[0]
    # A random sample across the whole file, not just the first 100 rows —
    # ids shouldn't all be small/contiguous the way df.head(100) would be.
    assert df["id"].max() > 200


def test_duckdb_sample_is_reproducible_given_same_seed():
    data = _csv_bytes(2000)
    df1, _ = _duckdb_sample_csv(_FakeUploadedFile(data), max_rows=50, random_state=7)
    df2, _ = _duckdb_sample_csv(_FakeUploadedFile(data), max_rows=50, random_state=7)
    assert sorted(df1["id"].tolist()) == sorted(df2["id"].tolist())


def test_duckdb_sample_returns_none_on_unparseable_content():
    garbage = _FakeUploadedFile(b"\x00\x01\x02\xff\xfe not a csv at all \x00")
    assert _duckdb_sample_csv(garbage, max_rows=100) is None


# ─────────────────────────────────────────────────────────────────────────
# load_data() — wiring: large files route through DuckDB, small files don't
# ─────────────────────────────────────────────────────────────────────────
def test_load_data_small_file_unaffected():
    f = _FakeUploadedFile(_csv_bytes(20))
    df, error, warnings = load_data(f)
    assert error is None
    assert len(df) == 20
    assert not any("DuckDB" in w for w in warnings)


def test_load_data_routes_large_file_through_duckdb():
    # Content is small, but a spoofed .size forces the large-file path —
    # isolates the routing decision from actually generating a huge file.
    f = _FakeUploadedFile(_csv_bytes(500), size=LARGE_FILE_THRESHOLD_BYTES + 1)
    df, error, warnings = load_data(f, max_rows=100)
    assert error is None
    assert len(df) == 100
    assert any("DuckDB" in w for w in warnings)


def test_load_data_falls_back_to_pandas_if_duckdb_cant_parse():
    # Spoofed large size, but content DuckDB's CSV reader chokes on (a
    # banner row it can't reconcile as well as pandas' dedicated recovery
    # path can) — must still succeed via the existing pandas fallback.
    messy = b"Company Sales Report\n\nid,value,label\n" + _csv_bytes(30).split(b"\n", 1)[1]
    f = _FakeUploadedFile(messy, size=LARGE_FILE_THRESHOLD_BYTES + 1)
    df, error, warnings = load_data(f)
    assert error is None
    assert len(df) >= 1


# ─────────────────────────────────────────────────────────────────────────
# check_sampling_fidelity — the self-verifying "did the sample turn out
# representative" check, shared by the CSV and Excel out-of-core paths.
# ─────────────────────────────────────────────────────────────────────────
def test_check_sampling_fidelity_no_stats_is_silent():
    df = pd.DataFrame({"a": [1, 2, 3]})
    assert check_sampling_fidelity(df, {"numeric": {}, "categorical": {}}) == []


def test_check_sampling_fidelity_numeric_within_threshold_reassures():
    df = pd.DataFrame({"value": [100.0, 101.0, 99.0, 100.0]})
    stats = {"numeric": {"value": {"mean": 100.0, "std": 5.0}}, "categorical": {}}
    result = check_sampling_fidelity(df, stats)
    assert len(result) == 1
    assert "✅" in result[0] and "1 column" in result[0]


def test_check_sampling_fidelity_numeric_drift_flags():
    # Sample mean (100) vs population mean (200) is a 50% relative gap —
    # well past the default 15% threshold.
    df = pd.DataFrame({"value": [99.0, 100.0, 101.0]})
    stats = {"numeric": {"value": {"mean": 200.0, "std": 5.0}}, "categorical": {}}
    result = check_sampling_fidelity(df, stats)
    assert len(result) == 1
    assert "⚠️" in result[0] and "value" in result[0]


def test_check_sampling_fidelity_numeric_zero_mean_uses_std_as_denominator():
    # pop_mean == 0 would divide-by-zero on a naive relative-error calc;
    # falls back to std as the denominator instead.
    df = pd.DataFrame({"value": [50.0, 51.0]})  # way off from a mean of 0
    stats = {"numeric": {"value": {"mean": 0.0, "std": 1.0}}, "categorical": {}}
    result = check_sampling_fidelity(df, stats)
    assert result and "⚠️" in result[0]


def test_check_sampling_fidelity_categorical_within_threshold_reassures():
    df = pd.DataFrame({"region": ["North"] * 33 + ["South"] * 67})
    stats = {"numeric": {}, "categorical": {"region": {"top_value": "South", "share": 0.65}}}
    result = check_sampling_fidelity(df, stats)
    assert len(result) == 1 and "✅" in result[0]


def test_check_sampling_fidelity_categorical_drift_flags():
    df = pd.DataFrame({"region": ["North"] * 90 + ["South"] * 10})
    stats = {"numeric": {}, "categorical": {"region": {"top_value": "South", "share": 0.60}}}
    result = check_sampling_fidelity(df, stats)
    assert len(result) == 1 and "⚠️" in result[0] and "region" in result[0]


def test_check_sampling_fidelity_mixed_only_flags_the_bad_column():
    df = pd.DataFrame({"good": [100.0, 100.0], "bad": [1.0, 1.0]})
    stats = {
        "numeric": {"good": {"mean": 100.0, "std": 2.0}, "bad": {"mean": 500.0, "std": 5.0}},
        "categorical": {},
    }
    result = check_sampling_fidelity(df, stats)
    assert len(result) == 1
    assert "bad" in result[0] and "good" not in result[0]


def test_check_sampling_fidelity_ignores_columns_missing_from_sample():
    df = pd.DataFrame({"a": [1.0, 2.0]})
    stats = {"numeric": {"a": {"mean": 1.5, "std": 0.5}, "gone": {"mean": 10.0, "std": 1.0}}, "categorical": {}}
    result = check_sampling_fidelity(df, stats)
    assert len(result) == 1 and "✅" in result[0]  # 'gone' silently skipped, 'a' passes


# ─────────────────────────────────────────────────────────────────────────
# CSV/DuckDB path — sampling now also runs the fidelity check
# ─────────────────────────────────────────────────────────────────────────
def test_duckdb_sample_over_cap_includes_fidelity_check():
    f = _FakeUploadedFile(_csv_bytes(2000))
    df, warnings = _duckdb_sample_csv(f, max_rows=100, random_state=42)
    assert any("Sampling fidelity" in w for w in warnings)


def test_duckdb_sample_under_cap_skips_fidelity_check():
    # No sampling happened (whole file fit), so there's nothing to verify.
    f = _FakeUploadedFile(_csv_bytes(50))
    df, warnings = _duckdb_sample_csv(f, max_rows=200)
    assert not any("Sampling fidelity" in w for w in warnings)


# ─────────────────────────────────────────────────────────────────────────
# _should_attempt_streaming_excel — the size/extension gate
# ─────────────────────────────────────────────────────────────────────────
pytest.importorskip("openpyxl")


def test_should_attempt_streaming_excel_false_for_small_file():
    f = _FakeUploadedFile(b"", name="data.xlsx", size=1_000)
    assert _should_attempt_streaming_excel(f) is False


def test_should_attempt_streaming_excel_true_above_threshold():
    f = _FakeUploadedFile(b"", name="data.xlsx", size=LARGE_FILE_THRESHOLD_BYTES + 1)
    assert _should_attempt_streaming_excel(f) is True


def test_should_attempt_streaming_excel_false_for_legacy_xls():
    # openpyxl can't stream the legacy binary format — always eager pandas.
    f = _FakeUploadedFile(b"", name="data.xls", size=LARGE_FILE_THRESHOLD_BYTES + 1)
    assert _should_attempt_streaming_excel(f) is False


def test_should_attempt_streaming_excel_false_when_size_unknown():
    f = _FakeUploadedFile(b"", name="data.xlsx")
    del f.size
    assert _should_attempt_streaming_excel(f) is False


# ─────────────────────────────────────────────────────────────────────────
# _stream_sample_excel — the actual streaming read
# ─────────────────────────────────────────────────────────────────────────
def _xlsx_bytes(n_rows: int, header: list | None = None) -> bytes:
    """Builds a minimal .xlsx in memory: id/value/label columns (mirrors
    _csv_bytes above) unless a custom header is given."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header if header is not None else ["id", "value", "label"])
    for i in range(n_rows):
        ws.append([i, i * 1.5, f"group{i % 3}"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_stream_sample_excel_returns_all_rows_when_under_cap():
    f = _FakeUploadedFile(_xlsx_bytes(50), name="data.xlsx")
    result = _stream_sample_excel(f, sheet_name=0, max_rows=200)
    assert result is not None
    df, warnings = result
    assert len(df) == 50
    assert list(df.columns) == ["id", "value", "label"]
    assert warnings == []


def test_stream_sample_excel_takes_random_sample_when_over_cap():
    f = _FakeUploadedFile(_xlsx_bytes(2000), name="data.xlsx")
    result = _stream_sample_excel(f, sheet_name=0, max_rows=100, random_state=42)
    assert result is not None
    df, warnings = result
    assert len(df) == 100
    assert any("2,000" in w for w in warnings)
    assert any("Sampling fidelity" in w for w in warnings)
    # A genuine reservoir sample across the whole sheet, not head(100).
    assert df["id"].max() > 200


def test_stream_sample_excel_is_reproducible_given_same_seed():
    data = _xlsx_bytes(2000)
    df1, _ = _stream_sample_excel(_FakeUploadedFile(data, name="data.xlsx"), sheet_name=0, max_rows=50, random_state=7)
    df2, _ = _stream_sample_excel(_FakeUploadedFile(data, name="data.xlsx"), sheet_name=0, max_rows=50, random_state=7)
    assert sorted(df1["id"].tolist()) == sorted(df2["id"].tolist())


def test_stream_sample_excel_bails_on_blank_header_row():
    f = _FakeUploadedFile(_xlsx_bytes(10, header=[None, None, None]), name="data.xlsx")
    assert _stream_sample_excel(f, sheet_name=0, max_rows=200) is None


def test_stream_sample_excel_returns_none_for_empty_sheet():
    import openpyxl

    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    f = _FakeUploadedFile(buf.getvalue(), name="data.xlsx")
    # A brand-new workbook's default sheet has no rows at all.
    assert _stream_sample_excel(f, sheet_name=0, max_rows=200) is None


# ─────────────────────────────────────────────────────────────────────────
# load_data() — wiring: large .xlsx files route through the streaming path
# ─────────────────────────────────────────────────────────────────────────
def test_load_data_routes_large_xlsx_through_streaming():
    f = _FakeUploadedFile(_xlsx_bytes(500), name="data.xlsx", size=LARGE_FILE_THRESHOLD_BYTES + 1)
    df, error, warnings = load_data(f, max_rows=100)
    assert error is None
    assert len(df) == 100
    assert any("Streamed" in w for w in warnings)


def test_load_data_falls_back_to_pandas_if_excel_streaming_bails():
    # Spoofed large size, but a blank header row the streaming path won't
    # touch — must still succeed via the existing eager openpyxl fallback
    # (with its own banner-row recovery).
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Company Sales Report"])
    ws.append(["id", "value", "label"])
    for i in range(10):
        ws.append([i, i * 1.5, f"group{i % 3}"])
    buf = io.BytesIO()
    wb.save(buf)
    f = _FakeUploadedFile(buf.getvalue(), name="data.xlsx", size=LARGE_FILE_THRESHOLD_BYTES + 1)
    df, error, warnings = load_data(f)
    assert error is None
    assert len(df) == 10
    assert list(df.columns) == ["id", "value", "label"]
