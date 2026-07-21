"""
Data Dictionary Generator — one AI-assisted pass documenting every column
(inferred description, dtype, example values, null %, and any quality
notes worth flagging), editable in the UI, then exportable as markdown or
a formatted .xlsx. Falls back to a templated (non-AI) description per
column when no Gemini key is configured, same fallback philosophy as the
rest of Prism.
"""

from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from modules.ai_analyst import build_data_context, call_gemini

_JSON_OBJECT_RE = re.compile(r"\{.*\}", re.DOTALL)

_PROMPT_TEMPLATE = (
    "{context}\n\n"
    "You are a data analyst writing a data dictionary for a colleague who has never seen "
    "this dataset. For EVERY column listed above, infer a one-sentence plain-English "
    "description of what it likely represents, based on its name, dtype, and sample values. "
    'Return a JSON object mapping column name to that description string, e.g. '
    '{{"customer_id": "A unique identifier for each customer."}}. Return ONLY the JSON '
    "object, no prose, no markdown code fences, one entry per column."
)


def _templated_description(col: str, ctype: str) -> str:
    label = col.replace("_", " ").replace("-", " ").strip()
    kind = {"numeric": "a numeric measure", "datetime": "a date/time value", "categorical": "a category label",
            "text": "free text", "all_null": "an empty column"}.get(ctype, "a value")
    return f"Likely {kind} named '{label}' — description auto-generated (no Gemini key configured)."


def generate_descriptions(model, df: pd.DataFrame, column_types: dict[str, str]) -> tuple[dict[str, str], Optional[str]]:
    """Ask Gemini for a one-line description per column. Returns
    ({column: description}, error) — error is None on success; on any
    failure every column still gets a templated description, never a gap.
    """
    fallback = {col: _templated_description(col, column_types.get(col, "text")) for col in df.columns}
    if model is None:
        return fallback, None

    context = build_data_context(df, column_types)
    text, error = call_gemini(model, _PROMPT_TEMPLATE.format(context=context))
    if error:
        return fallback, error

    match = _JSON_OBJECT_RE.search(text)
    if not match:
        return fallback, None
    try:
        parsed = json.loads(match.group(0))
    except json.JSONDecodeError:
        return fallback, None

    return {col: str(parsed.get(col) or fallback[col]) for col in df.columns}, None


def build_dictionary(
    df: pd.DataFrame, column_types: dict[str, str], quality_report: dict, descriptions: dict[str, str]
) -> list[dict]:
    """Assemble the full per-column dictionary rows: description, dtype,
    up to 3 example (non-null) values, null %, and a quality note (missing
    %, or "probable ID column" / "fully empty" flags).
    """
    rows = []
    missing_by_col = quality_report.get("missing_by_column", {})
    for col in df.columns:
        non_null = df[col].dropna()
        examples = ", ".join(str(v) for v in non_null.head(3).tolist()) or "—"
        missing_pct = missing_by_col.get(col, 0)

        notes = []
        ctype = column_types.get(col, "text")
        if ctype == "all_null":
            notes.append("fully empty")
        elif missing_pct >= 20:
            notes.append(f"{missing_pct}% missing — high")
        elif missing_pct > 0:
            notes.append(f"{missing_pct}% missing")
        if ctype != "all_null" and non_null.nunique() == len(df) and len(df) > 1:
            notes.append("all values unique — likely an ID column")

        rows.append({
            "Column": col,
            "Description": descriptions.get(col, ""),
            "Type": ctype,
            "Example Values": examples,
            "Missing %": missing_pct,
            "Notes": "; ".join(notes) or "—",
        })
    return rows


def to_markdown(rows: list[dict], dataset_name: str) -> str:
    lines = [f"# Data Dictionary — {dataset_name}", "", "| Column | Description | Type | Example Values | Missing % | Notes |",
              "|---|---|---|---|---|---|"]
    for r in rows:
        lines.append(
            f"| {r['Column']} | {r['Description']} | {r['Type']} | {r['Example Values']} | "
            f"{r['Missing %']}% | {r['Notes']} |"
        )
    return "\n".join(lines)


def to_xlsx_bytes(rows: list[dict], dataset_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"

    headers = ["Column", "Description", "Type", "Example Values", "Missing %", "Notes"]
    ws.append(headers)
    header_fill = PatternFill(start_color="0A0E17", end_color="0A0E17", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="00E5FF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for r in rows:
        ws.append([r["Column"], r["Description"], r["Type"], r["Example Values"], r["Missing %"], r["Notes"]])

    widths = [18, 46, 12, 36, 11, 30]
    for col_letter, width in zip("ABCDEF", widths):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
