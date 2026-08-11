"""
Data Engine — handles file ingestion, column type detection, and data quality
auditing for Prism. Every function here returns *new* objects rather than
mutating the DataFrame it's given, so the rest of the app can safely diff
before/after state.
"""

from __future__ import annotations

from typing import Optional, Union

import pandas as pd
import streamlit as st

# Rows beyond this threshold are truncated on load so the app stays responsive
# in the browser — Plotly and Streamlit both slow down heavily past ~50k rows.
MAX_ROWS = 50_000


def get_excel_sheet_names(uploaded_file) -> Optional[list[str]]:
    """Return the sheet names in an uploaded Excel file, or None if the file
    isn't Excel (or can't be read as one). Used to prompt the user to pick a
    sheet before loading a multi-sheet workbook — for a single-sheet workbook
    or a plain CSV, callers can skip the prompt and load directly.
    """
    filename = uploaded_file.name.lower()
    if not filename.endswith((".xlsx", ".xls")):
        return None
    try:
        uploaded_file.seek(0)
        sheet_names = pd.ExcelFile(uploaded_file).sheet_names
        uploaded_file.seek(0)
        return list(sheet_names)
    except Exception:
        return None


# Hard safety ceiling even when a caller asks for max_rows=None (Smart
# Sampling's "read the whole thing" path) — protects against a truly
# pathological upload blowing up server memory. Nothing between this and
# MAX_ROWS silently loses data; only above HARD_ROW_CEILING does Prism
# still truncate outright.
HARD_ROW_CEILING = 500_000

# Tried in order. utf-8-sig first (handles a BOM without leaving a stray
# character on the first header), then the two encodings that account for
# almost every "non-UTF-8" CSV in the wild — Windows exports (cp1252) and
# scraped/international datasets (latin-1, which never raises on any byte
# sequence, so it's the deliberate last resort rather than the first guess).
_ENCODING_FALLBACKS = ["utf-8-sig", "cp1252", "latin-1"]

# Tried in order. Comma first since it's overwhelmingly the common case and
# cheapest to confirm; semicolon covers most of continental Europe (whose
# own decimal comma pushes CSV exports to semicolons), then tab/pipe for
# database and log-style exports.
_DELIMITER_CANDIDATES = [",", ";", "\t", "|"]

# Above this many bytes, load_data() tries an out-of-core reader (DuckDB for
# CSV, streaming openpyxl for .xlsx — see below) before falling back to the
# pandas path above. Below it, pandas reading the whole file into memory
# first is fast enough that the extra machinery wouldn't pay for itself.
# 15 MB is comfortably above every sample dataset and typical Kaggle
# CSV/Excel export Prism has been tested against, so this only kicks in for
# genuinely large uploads.
LARGE_FILE_THRESHOLD_BYTES = 15 * 1024 * 1024

# How many columns a sampled-file's fidelity check inspects, at most — kept
# small so the extra population-stats query/pass stays cheap relative to the
# read itself. Excess columns are silently skipped past this many, but that
# skip is always surfaced in a warning (see _stream_sample_excel) rather than
# left implicit.
NUMERIC_FIDELITY_LIMIT = 8
CATEGORICAL_FIDELITY_LIMIT = 5


def _should_attempt_duckdb(uploaded_file) -> bool:
    """Whether load_data() should try the DuckDB path for this file.

    Only CSV-shaped files with a known size at/above the threshold qualify.
    .xlsx files above the threshold get their own streaming path instead
    (see _should_attempt_streaming_excel/_stream_sample_excel); .xls (the
    legacy binary format) and anything whose size can't be determined are
    treated conservatively — same behavior as today rather than risking a
    surprise.
    """
    size = getattr(uploaded_file, "size", None)
    if size is None:
        return False
    filename = getattr(uploaded_file, "name", "").lower()
    if filename.endswith((".xlsx", ".xls")):
        return False
    return size >= LARGE_FILE_THRESHOLD_BYTES


def _should_attempt_streaming_excel(uploaded_file) -> bool:
    """Whether load_data() should stream this .xlsx instead of reading it
    eagerly. The Excel analog of _should_attempt_duckdb above — same size
    gate, reusing LARGE_FILE_THRESHOLD_BYTES rather than a separate tuned
    constant. (.xlsx is zip-compressed, so on-disk size under-counts the
    decompressed row data — that only makes streaming kick in a bit later
    than it "ideally" would for spreadsheet-dense sheets, never earlier, so
    reusing the CSV threshold is a safe simplification, not a mistuning.)

    Only .xlsx qualifies — openpyxl can't read the legacy .xls binary format
    in a streaming/read-only fashion at all, so .xls always takes the eager
    pandas path further down.
    """
    size = getattr(uploaded_file, "size", None)
    if size is None or size < LARGE_FILE_THRESHOLD_BYTES:
        return False
    filename = getattr(uploaded_file, "name", "").lower()
    return filename.endswith(".xlsx")


def _duckdb_sample_csv(
    uploaded_file, max_rows: int, random_state: int = 42
) -> Optional[tuple[pd.DataFrame, list[str]]]:
    """Read a (potentially huge) CSV via DuckDB's out-of-core reader instead
    of pandas, returning at most `max_rows` rows.

    DuckDB streams the file rather than materializing it fully in Python
    memory, and — critically — when a sample is needed, it's a reservoir
    sample across the *entire* file rather than "keep the first max_rows
    rows" (pandas' `.head()` truncation, which silently over-represents
    whatever happens to be sorted or clustered near the top of the file,
    e.g. a dataset sorted by date or region).

    Returns (sampled_df, warnings) on success, or None on *any* failure —
    DuckDB not installed, a file it can't parse, an encoding it can't
    detect. This is a performance optimization on top of the pandas path,
    never a required capability, so callers should silently fall back to
    the existing pandas path rather than surface an error.
    """
    try:
        import duckdb
    except ImportError:
        return None

    import os
    import tempfile

    tmp_path = None
    try:
        uploaded_file.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name
        uploaded_file.seek(0)

        con = duckdb.connect(database=":memory:")
        try:
            total_rows = con.execute(
                "SELECT count(*) FROM read_csv_auto(?, ignore_errors=true)", [tmp_path]
            ).fetchone()[0]
            if total_rows == 0:
                return None
            if total_rows <= max_rows:
                df = con.execute("SELECT * FROM read_csv_auto(?, ignore_errors=true)", [tmp_path]).df()
                warnings = []
            else:
                df = con.execute(
                    "SELECT * FROM read_csv_auto(?, ignore_errors=true) "
                    f"USING SAMPLE {max_rows} ROWS (reservoir, {random_state})",
                    [tmp_path],
                ).df()
                warnings = [
                    f"This file has {total_rows:,} rows — too large to load in full. Used DuckDB's "
                    f"out-of-core reader to take a random {max_rows:,}-row sample across the *entire* "
                    f"file (not just the first {max_rows:,} rows), so the sample better represents the "
                    "whole dataset. The cleaned-data download will only reflect this sample."
                ]
            # DuckDB's auto-detection occasionally "succeeds" on malformed
            # input by degenerating to a near-empty parse (e.g. a banner row
            # mistaken for the real header, with every actual data row then
            # rejected as a width mismatch under ignore_errors=true) — a
            # technically-valid but useless DataFrame of all-null rows.
            # Pandas' dedicated banner/header-recovery heuristics (see
            # _recover_banner_row / _recover_missing_header above) handle
            # that case properly, so treat this as a failure and let the
            # caller fall back to them rather than accepting empty garbage.
            if df.dropna(how="all").empty:
                return None
            if total_rows > max_rows:
                # A sample was actually taken (as opposed to reading the
                # whole file) — worth checking whether it turned out
                # representative, since DuckDB already has the full file
                # open and this is cheap relative to the read itself.
                fidelity_stats = _population_stats_from_duckdb(con, tmp_path, df, total_rows)
                warnings.extend(check_sampling_fidelity(df, fidelity_stats))
            return df, warnings
        finally:
            con.close()
    except Exception:
        return None
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _population_stats_from_duckdb(con, tmp_path: str, df: pd.DataFrame, total_rows: int) -> dict:
    """Computes exact population statistics for a bounded subset of columns,
    straight from the full file via DuckDB SQL — no pandas ever holding the
    whole thing. Feeds check_sampling_fidelity() so a sample can be checked
    against the population it was drawn from, not just assumed good.

    Column *selection* (which ones qualify, and the LIMIT-based cap) is
    driven by the already-loaded sample's dtypes/cardinality — a proxy for
    the full file's shape, which is exactly the assumption a fidelity check
    is meant to catch if it's wrong. Any single column's query failing
    (a type DuckDB inferred differently than pandas did, an exotic value)
    is swallowed and that column is just skipped, since a partial fidelity
    check is still useful and one bad column shouldn't cancel the rest.
    """
    stats: dict = {"numeric": {}, "categorical": {}}
    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])][:NUMERIC_FIDELITY_LIMIT]
    categorical_cols = [
        c
        for c in df.columns
        if not pd.api.types.is_numeric_dtype(df[c]) and 1 < df[c].nunique(dropna=True) <= 50
    ][:CATEGORICAL_FIDELITY_LIMIT]

    if numeric_cols:
        select_clause = ", ".join(
            f'avg("{c}") AS "{c}__mean", stddev_pop("{c}") AS "{c}__std"' for c in numeric_cols
        )
        try:
            row = con.execute(
                f"SELECT {select_clause} FROM read_csv_auto(?, ignore_errors=true)", [tmp_path]
            ).fetchone()
            for i, c in enumerate(numeric_cols):
                mean, std = row[2 * i], row[2 * i + 1]
                if mean is not None:
                    stats["numeric"][c] = {"mean": float(mean), "std": float(std) if std is not None else 0.0}
        except Exception:
            pass

    for c in categorical_cols:
        try:
            top = con.execute(
                f'SELECT "{c}", count(*) FROM read_csv_auto(?, ignore_errors=true) '
                f'GROUP BY "{c}" ORDER BY count(*) DESC LIMIT 1',
                [tmp_path],
            ).fetchone()
            if top is not None and top[0] is not None:
                stats["categorical"][c] = {"top_value": top[0], "share": top[1] / total_rows}
        except Exception:
            continue

    return stats


def check_sampling_fidelity(sample_df: pd.DataFrame, population_stats: dict, rel_threshold: float = 0.15) -> list[str]:
    """Self-verifying check for an out-of-core sample: does the sample that
    actually got loaded still look like the full file it was drawn from?

    population_stats is {"numeric": {col: {"mean", "std"}}, "categorical":
    {col: {"top_value", "share"}}} — computed straight from the full file
    (see _population_stats_from_duckdb / _stream_sample_excel), never from
    pandas holding the whole thing. A column is flagged when the sample's
    own statistic drifts from the true population value by more than
    rel_threshold (15% by default — the same small/medium-effect spirit as
    the effect-size thresholds elsewhere in Prism, just applied to sampling
    error instead of a group difference). Returns a reassurance message
    instead of nothing when every checked column passed, so a clean check
    isn't indistinguishable from a check that never ran at all — the same
    "self-verifying agent" pattern Confounder Detection and Anomaly Drivers
    already use, applied at ingestion time instead of after analysis.
    """
    numeric_stats = population_stats.get("numeric", {})
    categorical_stats = population_stats.get("categorical", {})

    flagged: list[str] = []
    checked = 0
    for col, pop in numeric_stats.items():
        if col not in sample_df.columns:
            continue
        sample_series = pd.to_numeric(sample_df[col], errors="coerce").dropna()
        if sample_series.empty:
            continue
        checked += 1
        pop_mean, pop_std = pop["mean"], pop["std"]
        sample_mean = sample_series.mean()
        denom = abs(pop_mean) if abs(pop_mean) > 1e-9 else (pop_std if pop_std > 1e-9 else 1.0)
        rel_delta = abs(sample_mean - pop_mean) / denom
        if rel_delta > rel_threshold:
            flagged.append(
                f"⚠️ Sampling fidelity: column '{col}' — the sample's mean ({sample_mean:.3g}) differs from "
                f"the full file's mean ({pop_mean:.3g}) by {rel_delta:.0%}. Conclusions about this column "
                "may not hold for the full dataset."
            )

    for col, pop in categorical_stats.items():
        if col not in sample_df.columns:
            continue
        checked += 1
        top_value, pop_share = pop["top_value"], pop["share"]
        sample_share = (sample_df[col] == top_value).mean()
        delta = abs(sample_share - pop_share)
        if delta > rel_threshold:
            flagged.append(
                f"⚠️ Sampling fidelity: column '{col}' — its most common value in the full file "
                f"('{top_value}', {pop_share:.0%} of all rows) makes up {sample_share:.0%} of the sample, "
                f"a {delta:.0%}-point gap. Category proportions for this column may not hold for the full "
                "dataset."
            )

    if checked and not flagged:
        flagged.append(
            f"✅ Sampling fidelity check: {checked} column(s) checked against the full file's true "
            "statistics — the random sample looks representative."
        )
    return flagged


def _header_row_is_probably_data(df: pd.DataFrame) -> bool:
    """pandas' default header=0 treats a headerless CSV's first data row as
    column names — silently losing that row and mislabeling every column
    (e.g. UCI's adult.data, whose "headers" become "39", " State-gov",
    " 77516", ...). A dataset can easily be a realistic mix of numeric and
    categorical columns (age, workclass, fnlwgt, education, ...), so
    "how many header cells look numeric overall" is too weak a signal on
    its own. The precise version: look only at the columns pandas already
    inferred as numeric dtype from the body — a genuine header for a
    numeric column is a name like "age" or "fnlwgt", essentially never a
    number itself. If *every* numeric column's header also happens to look
    numeric, and there's more than one such column (ruling out the rare
    legitimate single numeric column name, e.g. a "2024" year column),
    that's a reliable sign the "header" is actually the first data row.
    """
    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if len(numeric_cols) < 2:
        return False

    def _looks_numeric(value) -> bool:
        try:
            float(str(value).strip())
            return True
        except ValueError:
            return False

    return all(_looks_numeric(c) for c in numeric_cols)


def _read_csv_with_encoding_fallback(uploaded_file) -> tuple[pd.DataFrame, str, str, bool]:
    """Read a CSV, sniffing encoding, delimiter, and header presence.

    Real Kaggle-style datasets (e.g. YouTube Trending's international video
    titles/tags) are routinely saved in cp1252/latin-1, not UTF-8 — a plain
    `pd.read_csv` throws UnicodeDecodeError on the first byte outside ASCII.
    Government open-data exports outside the US/UK (e.g. data.gouv.fr) are
    routinely semicolon-delimited, not comma — a plain comma parse either
    raises a ParserError partway through (once some row happens to contain
    an embedded comma) or "succeeds" with everything crammed into one column.
    Classic ML-repository datasets (e.g. UCI's adult.data) ship with no
    header row at all.

    Returns (dataframe, encoding_used, delimiter_used, header_recovered) so
    the caller can surface a warning when any fallback was needed.
    """
    last_error: Optional[Exception] = None
    fallback: Optional[tuple[pd.DataFrame, str, str]] = None
    for delimiter in _DELIMITER_CANDIDATES:
        for encoding in _ENCODING_FALLBACKS:
            uploaded_file.seek(0)
            try:
                df = pd.read_csv(uploaded_file, encoding=encoding, sep=delimiter)
            except UnicodeDecodeError as e:
                last_error = e
                continue
            except pd.errors.ParserError as e:
                last_error = e
                continue
            if df.shape[1] > 1:
                df, header_recovered = _recover_missing_header(df, uploaded_file, encoding, delimiter)
                return df, encoding, delimiter, header_recovered
            # Parsed cleanly but produced a single column — plausible if the
            # delimiter is simply wrong, but also possible the file
            # genuinely has one column. Keep it as a fallback and keep
            # looking for a delimiter that splits further.
            if fallback is None:
                fallback = (df, encoding, delimiter)
            break  # this delimiter's shape won't improve across encodings
    if fallback is not None:
        df, encoding, delimiter = fallback
        df, header_recovered = _recover_missing_header(df, uploaded_file, encoding, delimiter)
        return df, encoding, delimiter, header_recovered
    raise last_error


def _recover_missing_header(df: pd.DataFrame, uploaded_file, encoding: str, delimiter: str) -> tuple[pd.DataFrame, bool]:
    if not _header_row_is_probably_data(df):
        return df, False
    uploaded_file.seek(0)
    df = pd.read_csv(uploaded_file, encoding=encoding, sep=delimiter, header=None)
    df.columns = [f"Column_{i + 1}" for i in range(df.shape[1])]
    return df, True


def _unnamed_column_ratio(df: pd.DataFrame) -> float:
    if df.shape[1] == 0:
        return 0.0
    unnamed = sum(1 for c in df.columns if str(c).startswith("Unnamed:"))
    return unnamed / df.shape[1]


def _recover_banner_row(df: pd.DataFrame, reread) -> tuple[pd.DataFrame, bool]:
    """A banner/title row above the real header is a different failure mode
    from a missing header entirely (which _recover_missing_header already
    handles): a merged "Q3 Sales Report" cell, a blank row, THEN the real
    header. pandas reads that banner as the header, producing either a wall
    of "Unnamed: N" columns (multi-cell banner) or a bogus single-column
    parse (single-cell banner). `reread` is a zero-arg callable that
    re-parses the same source with header=1 (skip row 0, treat row 1 as the
    real header) — pulled out as a callable so this one recovery routine
    covers both the CSV and Excel branches in load_data(), which re-read
    from different underlying sources.
    """
    if not (_unnamed_column_ratio(df) > 0.5 or df.shape[1] <= 1) or df.shape[0] <= 1:
        return df, False
    try:
        retried = reread()
    except Exception:
        return df, False
    if retried.shape[1] > df.shape[1] or _unnamed_column_ratio(retried) < _unnamed_column_ratio(df):
        return retried, True
    return df, False


def _clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Strips stray whitespace from header cells — a very common artifact of
    hand-edited or copy-pasted spreadsheets ('Revenue ' vs 'Revenue' silently
    becoming two different columns downstream otherwise)."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _strip_object_column_whitespace(df: pd.DataFrame) -> pd.DataFrame:
    """Strips leading/trailing whitespace from string *values* (not just
    headers) — ' Bengaluru' and 'Bengaluru' silently becoming two different
    categories is a very common copy-paste/export artifact, and unlike
    numeric coercion this can't change what the data means, so it's safe to
    always apply rather than gate behind a confidence threshold."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]):
            df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)
    return df


def _stream_sample_excel(
    uploaded_file, sheet_name: Union[str, int], max_rows: int, random_state: int = 42
) -> Optional[tuple[pd.DataFrame, list[str]]]:
    """Reads a (potentially huge) .xlsx sheet by streaming rows via
    openpyxl's read_only mode, instead of pandas/openpyxl building the full
    in-memory workbook object graph first. The Excel analog of
    _duckdb_sample_csv above: takes a reservoir sample (Algorithm R) of at
    most max_rows rows in a single pass, and — in that same pass — tracks
    each numeric column's exact population mean/variance via Welford's
    online algorithm, so check_sampling_fidelity() can run without a second
    read of the file.

    Categorical-column fidelity isn't tracked here — unlike the CSV path,
    which gets it almost for free via a SQL GROUP BY, doing it safely in a
    single streaming pass needs a bounded-cardinality counter per column
    that gets abandoned once it overflows (a free-text column could
    otherwise grow one entry per row). That's a real follow-on, not
    attempted half-way; logged in the backlog instead.

    Returns (sampled_df, warnings) on success, or None on *any* failure —
    openpyxl not installed, an unreadable workbook, a header row that
    doesn't look usable (e.g. a banner row — the existing pandas fallback's
    banner-row recovery already handles that case properly, so this bails
    rather than trying to be equally clever). Same silent-fallback contract
    as _duckdb_sample_csv: a performance optimization, never a required
    capability.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    import random

    try:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(filename=uploaded_file, read_only=True, data_only=True)
        try:
            if isinstance(sheet_name, int):
                if sheet_name >= len(wb.worksheets):
                    return None
                ws = wb.worksheets[sheet_name]
            else:
                if sheet_name not in wb.sheetnames:
                    return None
                ws = wb[sheet_name]

            row_iter = ws.iter_rows(values_only=True)
            try:
                header = next(row_iter)
            except StopIteration:
                return None
            if not header or sum(1 for c in header if c is not None) < 2:
                return None
            columns = [str(c) if c is not None else f"Column_{i + 1}" for i, c in enumerate(header)]
            n_cols = len(columns)

            rng = random.Random(random_state)
            reservoir: list[tuple] = []
            total_rows = 0
            numeric_state: dict[int, list[float]] = {}  # col idx -> [count, mean, M2]
            NUMERIC_TRACK_LIMIT = 50

            for row in row_iter:
                total_rows += 1
                if len(reservoir) < max_rows:
                    reservoir.append(row)
                else:
                    j = rng.randint(0, total_rows - 1)
                    if j < max_rows:
                        reservoir[j] = row
                for idx in range(min(n_cols, NUMERIC_TRACK_LIMIT, len(row))):
                    val = row[idx]
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        state = numeric_state.get(idx)
                        if state is None:
                            numeric_state[idx] = [1, float(val), 0.0]
                        else:
                            state[0] += 1
                            delta = val - state[1]
                            state[1] += delta / state[0]
                            state[2] += delta * (val - state[1])
        finally:
            wb.close()
    except Exception:
        return None

    if total_rows == 0:
        return None

    df = pd.DataFrame(reservoir, columns=columns)
    warnings: list[str] = []
    if total_rows > max_rows:
        warnings.append(
            f"This Excel sheet has {total_rows:,} rows — too large to load in full. Streamed it row-by-row "
            f"(without building the full sheet in memory) and took a random {max_rows:,}-row sample across "
            "the entire sheet, not just the top. The cleaned-data download will only reflect this sample."
        )
        if n_cols > NUMERIC_TRACK_LIMIT:
            warnings.append(f"(Sampling fidelity check limited to the first {NUMERIC_TRACK_LIMIT} of {n_cols} columns.)")
        population_stats = {
            "numeric": {
                columns[idx]: {"mean": mean, "std": (m2 / (count - 1)) ** 0.5}
                for idx, (count, mean, m2) in numeric_state.items()
                if count >= 2
            },
            "categorical": {},
        }
        warnings.extend(check_sampling_fidelity(df, population_stats))

    return df, warnings


def load_data(
    uploaded_file, sheet_name: Union[str, int] = 0, max_rows: Optional[int] = MAX_ROWS
) -> tuple[Optional[pd.DataFrame], Optional[str], list[str]]:
    """Read an uploaded CSV/Excel file into a DataFrame.

    sheet_name is only used for Excel files — defaults to the first sheet
    (matching plain pd.read_excel's default) so existing single-sheet callers
    are unaffected. Pass an explicit sheet name once the user has picked one
    from get_excel_sheet_names().

    max_rows=MAX_ROWS (default) preserves the original behavior: silently
    truncate to the first MAX_ROWS rows with a warning. Pass max_rows=None
    for Smart Sampling's flow, which needs the (near-)full frame so it can
    offer a *chosen* random/stratified sample instead of always "the first
    N rows" — still hard-capped at HARD_ROW_CEILING either way.

    Returns a (dataframe, error_message, warnings) tuple. On failure,
    dataframe is None and error_message explains why. On success,
    error_message is None and dataframe is ready to use.
    """
    warnings: list[str] = []
    filename = uploaded_file.name.lower()
    effective_cap = max_rows if max_rows is not None else HARD_ROW_CEILING

    # Large-file fast path: DuckDB reads the file out-of-core and samples
    # directly, so pandas never has to materialize the whole thing just to
    # truncate it a moment later. Only attempted above LARGE_FILE_THRESHOLD_
    # BYTES (see _should_attempt_duckdb) and only for CSV-shaped files; on
    # any failure (DuckDB missing, a parse quirk it can't recover from the
    # way the pandas fallback below can) this silently falls through to the
    # existing behavior rather than surfacing an error.
    duckdb_result = _duckdb_sample_csv(uploaded_file, effective_cap) if _should_attempt_duckdb(uploaded_file) else None
    excel_stream_result = (
        _stream_sample_excel(uploaded_file, sheet_name, effective_cap)
        if duckdb_result is None and _should_attempt_streaming_excel(uploaded_file)
        else None
    )

    if duckdb_result is not None:
        df, warnings = duckdb_result
    elif excel_stream_result is not None:
        df, warnings = excel_stream_result
    else:
        try:
            if filename.endswith((".xlsx", ".xls")):
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name)

                def _reread_excel():
                    uploaded_file.seek(0)
                    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=1)

                df, banner_skipped = _recover_banner_row(df, _reread_excel)
                if banner_skipped:
                    warnings.append("Detected a banner row above the real header and skipped it.")
            else:
                # Everything else — .csv, .txt, .dat, .data, or no extension at
                # all — is treated as a delimited-text candidate rather than
                # rejected by name. Real downloads routinely arrive without a
                # clean .csv extension (government portals with extensionless
                # API URLs, UCI's *.data files); the parser below still raises a
                # clear, specific error if the content genuinely isn't
                # delimited text, so nothing is silently accepted.
                df, encoding_used, delimiter_used, header_recovered = _read_csv_with_encoding_fallback(uploaded_file)
                if encoding_used != "utf-8":
                    warnings.append(
                        f"This file wasn't valid UTF-8 (common for datasets scraped or exported outside "
                        f"the US/UK — e.g. non-English titles or tags). Read successfully using "
                        f"{encoding_used} instead."
                    )
                if delimiter_used != ",":
                    shown = {"\t": "tab"}.get(delimiter_used, delimiter_used)
                    warnings.append(
                        f"This file uses '{shown}' instead of a comma to separate columns (common outside "
                        f"the US/UK, or in database/log exports) — detected automatically."
                    )
                if header_recovered:
                    warnings.append(
                        "This file doesn't appear to have a header row (the first row looked like data, "
                        "not column names), so it would otherwise have been silently used as one — Prism "
                        "generated generic column names (Column_1, Column_2, ...) and kept every row."
                    )
                else:

                    def _reread_csv():
                        uploaded_file.seek(0)
                        return pd.read_csv(uploaded_file, encoding=encoding_used, sep=delimiter_used, header=1)

                    df, banner_skipped = _recover_banner_row(df, _reread_csv)
                    if banner_skipped:
                        warnings.append("Detected a banner row above the real header and skipped it.")
        except pd.errors.EmptyDataError:
            return None, "The uploaded file is empty.", warnings
        except pd.errors.ParserError as e:
            return None, f"Could not parse the file — it may be malformed. Details: {e}", warnings
        except UnicodeDecodeError:
            return None, "Could not decode the file in UTF-8, CP1252, or Latin-1. Try re-saving it as UTF-8 CSV.", warnings
        except ValueError as e:
            return None, f"Could not read the file. Details: {e}", warnings
        except Exception as e:  # last-resort catch so a bad upload never hard-crashes the app
            return None, f"Unexpected error while reading the file: {e}", warnings

    if df.shape[0] == 0:
        return None, "The uploaded file contains no rows.", warnings
    if df.shape[1] == 0:
        return None, "The uploaded file contains no columns.", warnings

    df = _clean_column_names(df)
    df = _strip_object_column_whitespace(df)

    # Some Excel exports leave fully-empty trailing rows behind — drop them before
    # they skew the quality report. Fully-empty *columns* are intentionally kept:
    # they're a data-quality signal (flagged in the report, droppable in Cleaning
    # Controls) rather than something to silently discard on load.
    df = df.dropna(how="all")

    if df.shape[0] > effective_cap:
        if max_rows is None:
            warnings.append(
                f"The file has {df.shape[0]:,} rows — even Smart Sampling's full read caps at "
                f"{HARD_ROW_CEILING:,} rows to protect memory. Truncated to the first {HARD_ROW_CEILING:,}."
            )
        else:
            warnings.append(
                f"The file has {df.shape[0]:,} rows — Prism sampled the first {MAX_ROWS:,} "
                "rows to keep the app responsive. The cleaned-data download will only "
                "reflect this sample."
            )
        df = df.head(effective_cap).reset_index(drop=True)

    return df, None, warnings


def sample_dataframe(
    df: pd.DataFrame, method: str, target_rows: int, strat_col: Optional[str] = None
) -> tuple[pd.DataFrame, str]:
    """Smart Sampling: reduce a large DataFrame to ~target_rows.

    method="random": a uniform random sample.
    method="stratified": sampled proportionally within each group of
    strat_col, so category shares in the sample match the full dataset
    (e.g. if 30% of rows are "West" region in the full data, ~30% of the
    sampled rows are too) — reproducible via a fixed random_state.

    Returns (sampled_df, explanation) — explanation is the one-line reason
    shown in the persistent sampling banner.
    """
    n_total = len(df)
    if n_total <= target_rows:
        return df, ""

    if method == "stratified" and strat_col and strat_col in df.columns:
        frac = target_rows / n_total
        sampled = (
            df.groupby(strat_col, group_keys=False)[df.columns.tolist()]
            .apply(lambda g: g.sample(frac=frac, random_state=42))
            .reset_index(drop=True)
        )
        explanation = (
            f"Working on a {len(sampled):,}-row stratified sample of {n_total:,} rows "
            f"(proportions by '{strat_col}' preserved — each category keeps its original share)."
        )
    else:
        sampled = df.sample(n=target_rows, random_state=42).reset_index(drop=True)
        explanation = f"Working on a {len(sampled):,}-row random sample of {n_total:,} rows."

    return sampled, explanation


def _looks_like_datetime(series: pd.Series) -> bool:
    """Heuristic: try parsing an object column as dates and see how well it sticks."""
    non_null = series.dropna()
    if non_null.empty:
        return False
    # Pure numbers (e.g. "42") shouldn't be treated as dates even though pandas
    # can sometimes coerce them into odd epoch-based timestamps.
    if pd.to_numeric(non_null, errors="coerce").notna().mean() > 0.9:
        return False
    parsed = pd.to_datetime(non_null, errors="coerce", format="mixed")
    return parsed.notna().mean() > 0.9


@st.cache_data(show_spinner=False)
def detect_column_types(df: pd.DataFrame) -> dict[str, str]:
    """Classify each column as one of: numeric, datetime, categorical, text, all_null.

    - numeric / datetime are inferred from the pandas dtype, or by attempting
      to parse object columns as dates.
    - Among remaining object columns, low-cardinality ones are 'categorical'
      (good for pie/bar charts) and high-cardinality free text is 'text'.

    @st.cache_data: this is called 25+ times across app.py, and Streamlit
    reruns the whole script on every widget interaction anywhere in the app
    — without caching, a dataset's column types were being re-inferred from
    scratch dozens of times per session even when nothing about the data had
    changed. Pure function of `df`'s contents, so caching is exact, not an
    approximation: a genuinely different `df` is a cache miss, not a stale hit.
    """
    column_types: dict[str, str] = {}

    for col in df.columns:
        series = df[col]

        if series.isna().all():
            column_types[col] = "all_null"
            continue

        if pd.api.types.is_bool_dtype(series):
            column_types[col] = "categorical"
        elif pd.api.types.is_numeric_dtype(series):
            column_types[col] = "numeric"
        elif pd.api.types.is_datetime64_any_dtype(series):
            column_types[col] = "datetime"
        elif _looks_like_datetime(series):
            column_types[col] = "datetime"
        else:
            non_null = series.dropna()
            n_unique = non_null.nunique()
            unique_ratio = n_unique / len(non_null) if len(non_null) else 0
            if n_unique <= 50 and unique_ratio < 0.5:
                column_types[col] = "categorical"
            else:
                column_types[col] = "text"

    return column_types


def _format_bytes(num_bytes: float) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if num_bytes < 1024:
            return f"{num_bytes:.1f} {unit}"
        num_bytes /= 1024
    return f"{num_bytes:.1f} TB"


def detect_outliers_iqr(series: pd.Series) -> tuple[int, float]:
    """Count outliers in a numeric series using the IQR (Tukey's fence) method."""
    clean = series.dropna()
    if len(clean) < 4:
        return 0, 0.0
    q1, q3 = clean.quantile(0.25), clean.quantile(0.75)
    iqr = q3 - q1
    lower, upper = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    outliers = clean[(clean < lower) | (clean > upper)]
    return len(outliers), round(100 * len(outliers) / len(clean), 2)


@st.cache_data(show_spinner=False)
def get_data_quality_report(df: pd.DataFrame, column_types: dict[str, str]) -> dict:
    """Build a single dict summarizing dataset health for the Overview tab and report export.

    @st.cache_data: same rationale as detect_column_types above — this scans
    every column for missing values and IQR outliers, and was being
    recomputed from scratch on every rerun at 10+ call sites even when the
    underlying data hadn't changed since the last run.
    """
    n_rows, n_cols = df.shape
    missing_by_column = {col: round(100 * df[col].isna().sum() / n_rows, 2) for col in df.columns}

    outliers = {}
    for col, ctype in column_types.items():
        if ctype == "numeric":
            count, pct = detect_outliers_iqr(df[col])
            outliers[col] = {"count": count, "pct": pct}

    total_cells = n_rows * n_cols
    return {
        "n_rows": n_rows,
        "n_cols": n_cols,
        "missing_by_column": missing_by_column,
        "total_missing_cells": int(df.isna().sum().sum()),
        "total_missing_pct": round(100 * df.isna().sum().sum() / total_cells, 2) if total_cells else 0.0,
        "duplicate_rows": int(df.duplicated().sum()),
        "memory_usage": _format_bytes(df.memory_usage(deep=True).sum()),
        "outliers": outliers,
        "all_null_columns": [c for c, t in column_types.items() if t == "all_null"],
    }


HEALTH_COMPONENT_WEIGHTS = {
    "completeness": 30, "consistency": 25, "uniqueness": 15, "validity": 15, "outlier_burden": 15,
}


def get_health_breakdown(
    quality_report: dict, column_types: dict[str, str], pii_findings: Optional[dict] = None
) -> dict:
    """The explainable 0-100 Data Health Score, broken into its 5 weighted
    components (see HEALTH_COMPONENT_WEIGHTS) so the number is never just
    "magic" — every component is independently visible in the Overview
    tab's breakdown expander and the sticky header's mini-badge.

        completeness (30)   — share of non-missing cells
        consistency  (25)   — share of columns still stuck as free 'text'
                               instead of a proper type (numeric/date/etc.)
        uniqueness   (15)   — share of rows that aren't exact duplicates
        validity     (15)   — unmasked PII exposure + fully-empty columns
        outlier_burden (15) — average IQR-outlier share across numeric columns
    """
    n_rows = quality_report["n_rows"] or 1
    n_cols = quality_report["n_cols"] or 1

    completeness = 30 * (1 - quality_report["total_missing_pct"] / 100)

    text_cols = sum(1 for t in column_types.values() if t == "text")
    consistency = 25 * (1 - 0.5 * (text_cols / n_cols))

    duplicate_pct = 100 * quality_report["duplicate_rows"] / n_rows
    uniqueness = 15 * (1 - duplicate_pct / 100)

    validity = 15.0
    if pii_findings:
        from modules import pii_detector
        if pii_detector.has_findings(pii_findings):
            validity -= 7
    validity -= min(len(quality_report["all_null_columns"]) * 3, 8)

    outlier_pcts = [info["pct"] for info in quality_report["outliers"].values()]
    avg_outlier_pct = (sum(outlier_pcts) / len(outlier_pcts)) if outlier_pcts else 0
    outlier_burden = 15 * (1 - min(avg_outlier_pct, 30) / 30)

    components = {
        "completeness": round(max(0, min(30, completeness))),
        "consistency": round(max(0, min(25, consistency))),
        "uniqueness": round(max(0, min(15, uniqueness))),
        "validity": round(max(0, min(15, validity))),
        "outlier_burden": round(max(0, min(15, outlier_burden))),
    }
    components["total"] = max(0, min(100, sum(components.values())))
    return components


def get_health_score(
    quality_report: dict, column_types: Optional[dict[str, str]] = None, pii_findings: Optional[dict] = None
) -> int:
    """The single 0-100 number — sticky header badge, Overview's gauge, and
    every before/after delta. column_types defaults to empty for callers
    that haven't been updated (consistency component reads as unaffected
    rather than erroring), but every in-app call site passes it.
    """
    return get_health_breakdown(quality_report, column_types or {}, pii_findings)["total"]
